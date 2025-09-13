from flask import Flask, request, render_template_string
from openpyxl import load_workbook

app = Flask(__name__)

# Path to your Excel file inside your project folder
file_path = "2018 Meskel Promises.xlsx"

# Load Excel data into a list of dictionaries
def load_excel_data():
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    headers = [cell.value for cell in sheet[1]]  # first row = column names
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data

data = load_excel_data()

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Meskel Promise Finder 🎉🔥</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/canvas-confetti@1.6.0/dist/confetti.browser.min.js"></script>
    <style>
        html, body { height: 100%; margin: 0; }
        body {
            background-image: url('https://upload.wikimedia.org/wikipedia/commons/2/2e/Ethiopia_Adew_Flower_Background.jpg');
            background-size: cover;
            background-position: center;
            display: flex;
            align-items: center;
            justify-content: center;
            position: relative;
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        }
        .container {
            background: white;
            border-radius: 15px;
            padding: 40px 30px;
            max-width: 600px;
            width: 90%;
            text-align: center;
            box-shadow: 0px 8px 20px rgba(0,0,0,0.3);
            position: relative;
            z-index: 1;
        }
        h1 { color: #d84315; font-weight: bold; margin-bottom: 20px; }
        .btn-custom { background-color: #d84315; color: white; font-weight: bold; }
        .btn-custom:hover { background-color: #bf360c; }
        .promise-box { margin-top: 20px; padding: 20px; background: #fff3e0; border-radius: 10px; border-left: 6px solid #d84315; font-size: 18px; }
        .top-images { position: absolute; top: -50px; width: 100%; display: flex; justify-content: space-between; padding: 0 30px; z-index: 2; }
        .top-images img { width: 60px; height: 60px; animation: float 3s ease-in-out infinite alternate; }
        @keyframes float { 0% { transform: translateY(0); } 50% { transform: translateY(-10px); } 100% { transform: translateY(0); } }
        .floating-flower { position: fixed; width: 50px; height: 50px; background-image: url('https://upload.wikimedia.org/wikipedia/commons/3/33/Adey_Flower.jpg'); background-size: cover; pointer-events: none; animation: floatUp 12s linear infinite; }
        @keyframes floatUp { 0% { transform: translateY(100vh) translateX(0); opacity: 1; } 100% { transform: translateY(-50px) translateX(50px); opacity: 0.5; } }
    </style>
</head>
<body>
    <div class="top-images">
        <img src="https://upload.wikimedia.org/wikipedia/commons/8/88/Christian_cross.svg" alt="Jesus Cross Left">
        <img src="https://upload.wikimedia.org/wikipedia/commons/8/88/Christian_cross.svg" alt="Jesus Cross Right">
    </div>

    <div class="floating-flower" style="left:5%; animation-delay:0s;"></div>
    <div class="floating-flower" style="left:25%; animation-delay:3s;"></div>
    <div class="floating-flower" style="left:55%; animation-delay:5s;"></div>
    <div class="floating-flower" style="left:75%; animation-delay:2s;"></div>

    <div class="container">
        <h1>🌼 Meskel Promises 🔥</h1>
        <p>Enter your details to see your secret promise for 2018 E.C.</p>
        <form method="POST">
            <div class="mb-3">
                <input type="text" class="form-control" name="name" placeholder="Enter your Name" required>
            </div>
            <div class="mb-3">
                <input type="text" class="form-control" name="phone" placeholder="Enter your Phone Number" required>
            </div>
            <button type="submit" class="btn btn-custom">Show My Promise</button>
        </form>
        {% if promise %}
        <div class="promise-box" id="promise-box">
            <strong>Dear {{ name }},</strong><br>
            Your promise for the 2018 Meskel Festival is:<br><br>
            "{{ promise }}"
        </div>
        <script>
            confetti({ particleCount: 150, spread: 70, origin: { y: 0.6 } });
        </script>
        {% elif error %}
        <div class="alert alert-danger mt-3">{{ error }}</div>
        {% endif %}
    </div>
</body>
</html>
"""

@app.route("/", methods=["GET", "POST"])
def home():
    promise = None
    error = None
    name = ""
    if request.method == "POST":
        name = request.form.get("name").strip()
        phone = request.form.get("phone").strip()
        # Search in data loaded from Excel
        match = [row for row in data if str(row['Name']).strip().lower() == name.lower() and str(row['Phone']).strip() == phone]
        if match:
            promise = match[0]['Promises']
        else:
            error = "No matching record found. Please check your name and phone."
    return render_template_string(HTML_TEMPLATE, promise=promise, error=error, name=name)

import os

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))  # Use Render's assigned port or 5000 locally
    app.run(host="0.0.0.0", port=port, debug=True)
